import io
from datetime import datetime, date, time
from typing import Optional

import openpyxl

# El archivo DICKSON tiene los headers en la fila 7 (1-based)
_HEADER_ROW = 7


def parse_dickson_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Parsea un archivo Excel raw del datalogger DICKSON PR1000.
    Retorna lista de dicts: {timestamp, presion_mca, temperatura_c}
    La presión ya viene en mH2O (equivalente a mca), no requiere conversión.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # Buscar la hoja de datos que tenga 'Fecha' en la fila 7
    # Se saltean chart sheets (no tienen método .cell)
    ws = None
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        if not hasattr(candidate, "cell"):
            continue
        for c in range(1, 10):
            col_val = candidate.cell(row=_HEADER_ROW, column=c).value
            if col_val and "echa" in str(col_val).lower():
                ws = candidate
                break
        if ws is not None:
            break

    if ws is None:
        raise ValueError(
            "No se encontró ninguna hoja con columna 'Fecha' en la fila 7. "
            "Verifique que el archivo sea exportado directamente del datalogger DICKSON."
        )

    fecha_col = tiempo_col = temp_col = presion_col = None
    readings = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == _HEADER_ROW:
            for col_idx, val in enumerate(row):
                if val is None:
                    continue
                h = str(val).lower()
                if "echa" in h and fecha_col is None:
                    fecha_col = col_idx
                elif "iempo" in h and tiempo_col is None:
                    tiempo_col = col_idx
                elif "emp" in h and temp_col is None:
                    temp_col = col_idx
                elif "res" in h and presion_col is None:
                    presion_col = col_idx

            if fecha_col is None or presion_col is None:
                raise ValueError(
                    "Formato no reconocido: no se encontraron columnas 'Fecha' o 'Presión' en la fila 7."
                )

        elif row_idx > _HEADER_ROW:
            if not row or row[fecha_col] is None:
                continue

            presion = row[presion_col]
            if presion is None:
                continue

            dt = _combine_datetime(
                row[fecha_col],
                row[tiempo_col] if tiempo_col is not None else None,
            )
            if dt is None:
                continue

            temperatura = row[temp_col] if temp_col is not None else None

            readings.append({
                "timestamp": dt,
                "presion_mca": round(float(presion), 4),
                "temperatura_c": round(float(temperatura), 2) if temperatura is not None else None,
            })

    wb.close()
    return readings


def _combine_datetime(fecha, tiempo) -> Optional[datetime]:
    try:
        if isinstance(fecha, datetime):
            d = fecha.date()
        elif isinstance(fecha, date):
            d = fecha
        else:
            return None

        if tiempo is None:
            return datetime(d.year, d.month, d.day)
        elif isinstance(tiempo, datetime):
            t = tiempo.time()
        elif isinstance(tiempo, time):
            t = tiempo
        else:
            return None

        return datetime.combine(d, t)
    except Exception:
        return None
